from openpyxl import styles
from openpyxl.styles import Font, Border, Alignment, Color, Side,PatternFill, NamedStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from operator import itemgetter
from datetime import date,datetime
import random, csv
from Function.FuncMatrix import Func_PY_Matrix as FuncMatrix

# FUNZIONE PER STILI
def SetStyleExcel(nStyle,FontSize,Bold,hAlign,vAlign,cStart,cEnd,cFillType,bStyle,cBorder,sBorder):
    nStyle.font=styles.Font(size=FontSize,bold=Bold)
    nStyle.alignment=styles.alignment.Alignment(horizontal=hAlign,vertical=vAlign)
    nStyle.fill=styles.fills.PatternFill(start_color=cStart, end_color=cEnd, fill_type=cFillType)
    nStyle.border=styles.borders.Border(left=Side(style=bStyle,color=cBorder,border_style=sBorder),
                                            right=Side(style=bStyle,color=cBorder,border_style=sBorder),
                                            top=Side(style=bStyle,color=cBorder,border_style=sBorder),
                                            bottom=Side(style=bStyle,color=cBorder,border_style=sBorder))
    return nStyle

def SetStyleExcel_Font(nStyle,FontSize,Bold):
    nStyle.font=styles.Font(size=FontSize,bold=Bold)
    return nStyle

def SetStyleExcel_Alignement(nStyle,hAlign,vAlign):
    nStyle.alignment=styles.alignment.Alignment(horizontal=hAlign,vertical=vAlign)
    return nStyle

def SetStyleExcel_Rotation(nStyle,rotation):
    nStyle.alignment=styles.alignment.Alignment(textRotation=rotation)
    return nStyle

def SetStyleExcel_PatternFill(nStyle,cStart,cEnd,cFillType):
    nStyle.fill=styles.fills.PatternFill(start_color=cStart, end_color=cEnd, fill_type=cFillType)
    return nStyle

def SetStyleExcel_Border (nStyle,bStyle,cBorder,sBorder):
    nStyle.border=styles.borders.Border(left=Side(style=bStyle,color=cBorder,border_style=sBorder),
                                            right=Side(style=bStyle,color=cBorder,border_style=sBorder),
                                            top=Side(style=bStyle,color=cBorder,border_style=sBorder),
                                            bottom=Side(style=bStyle,color=cBorder,border_style=sBorder))
    return nStyle

def SetStyleCell(sheet, rowStart,rowEnd, colStart,colEnd,typeStyle):
    for colNum in range(colStart,colEnd+1,1):
        for rowNum in range (rowStart,rowEnd+1,1):
            sheet.cell(row=rowNum, column=colNum).style=typeStyle

#FUNZIONI PER SET BORDI EXCEL
def SetBorderAll(sheet, rowStart,rowEnd, colStart,colEnd,typeBorder):
    black='FF000000'
    Select_Border=Border(left=Side(style=typeBorder,color=black,border_style=typeBorder),
                         right=Side(style=typeBorder,color=black,border_style=typeBorder),
                         top=Side(style=typeBorder,color=black,border_style=typeBorder),
                         bottom=Side(style=typeBorder,color=black,border_style=typeBorder))
    for colNum in range(colStart,colEnd+1,1):
        for rowNum in range (rowStart,rowEnd+1,1):
            sheet.cell(row=rowNum, column=colNum).border=Select_Border

def SetBorderRight(sheet, rowStart,rowEnd, colStart,colEnd,typeBorder):
    black='FF000000'
    Select_Border=Border(right=Side(style=typeBorder,color=black,border_style=typeBorder))
    for colNum in range(colStart,colEnd+1,1):
        for rowNum in range (rowStart,rowEnd+1,1):
            sheet.cell(row=rowNum, column=colNum).border=Select_Border

# FUNZIONI COLORI
def createRandomColor(startColor,endColor):
    r=str(hex(random.randrange(startColor,endColor)))
    g=str(hex(random.randrange(startColor,endColor)))
    b=str(hex(random.randrange(startColor,endColor)))
    colorRand='FF'+r[2:]+g[2:]+b[2:]
    return colorRand

def RenameActiveSheet(wb,sh_name):
    act_sheet=wb.active
    act_sheet.title=sh_name
    return act_sheet

# FUNZIONI COUNT - SUM
def SumDevice(ws):
    sum=0
    col_ini=2
    i=2
    rng_col=CreateRangeCell(ws,col_ini,ws.max_column,3,3)
    for col in rng_col:
        rng_row=CreateRangeCell(ws,i,i,5,ws.max_row)
        for row in rng_row:
            if(row.value):
                sum=sum+row.value
                ws.cell(row=4,column=i, value=sum)
        i=i+1
        sum=0

def countDevice(ws,arrRis,arrayStateSelected):
    #Set Range for search Location
    rng_location=CreateRangeCell(ws,1,1,5,ws.max_row)
    #Set Range for search Model
    rng_model=CreateRangeCell(ws,2,ws.max_column,3,3)
    count=0
    for x in arrRis:
        locationName=arrRis[count][2]
        stateName=arrRis[count][5]
        modelName=arrRis[count][11]
        vendorName=arrRis[count][10]
        count=count+1
        #Search row of location and column for Model
        row_location=SearchRow(rng_location,locationName)
        col_model=SearchColumn(rng_model,modelName)
        #Verify if the state is selected by user - Return 1 if is selected, 0 if is not selected
        stateSelect=FuncMatrix.SearchElemArrayNew(stateName,arrayStateSelected)
        if stateSelect==1:
            #Add 1 to cell found
            ndev=ws.cell(row=row_location, column=col_model).value
            if(ndev is None):
                ndev=0
            ws.cell(row=row_location, column=col_model).value=ndev+1

# FUNZIONI SEARCH

def SearchRow(rng,elem):
    for x in rng:
        if(x.value==elem):
            row_select=x.row
    return row_select

def SearchColumn(rng,elem):
    for x in rng:
        if(x.value==elem):
            col_select=x.column
    return col_select

#FUNZIONI - CREA FOGLI

def createSheetManu(wb,matModels,matLocations,arrTot,sTitle1,sTitle2,sType,sModel,sTotali,sLocation,sText):
    
    sheetManu=RenameActiveSheet(wb,'Manutenzioni')
    
    lModels=len(matModels)
    entity_name=arrTot[0][13]
    data1=datetime.now().date()
    
    #Fixed Text and Cells
    sheetManu.merge_cells(start_row=1,start_column=1,end_row=1,end_column=lModels+1)
    SetStyleCell(sheetManu,1,1,1,1,sTitle1)

    sheetManu.cell(row=1, column=1, value='MANUTENZIONE '+ entity_name + ' ' + str(data1.day) + '-' + str(data1.month) + '-' + str(data1.year))
    sheetManu.cell(row=2, column=1, value='Sedi')
    sheetManu.cell(row=3, column=1, value='Negozi')
    sheetManu.cell(row=4, column=1, value='Totali')
    SetStyleCell(sheetManu,2,4,1,1,sTitle2)

    #Cycle to copy Model and Type in Excel file
    count=0
    matModels.sort(key=itemgetter(0)) #Sort Matrix by Type
    current_type=""
    startColumnTypes=2
    for x in matModels:
        #Generate casual color and use of a cell
        colorRandom=createRandomColor(150,255)
        sModel.fill=styles.fills.PatternFill(start_color=colorRandom, end_color=colorRandom, fill_type='solid')
        #Cycle to write only 1 time the type
        if current_type=="":                        #First element
            current_type=matModels[count][0]
        if (current_type!=matModels[count][0]):     #Current type is != from the previous
            sType.fill=styles.fills.PatternFill(start_color=colorRandom, end_color=colorRandom, fill_type='solid')
            sheetManu.merge_cells(start_row=2,start_column=startColumnTypes,end_row=2,end_column=count+1)
            sheetManu.cell(row=2, column=startColumnTypes, value=current_type).style=sType
            SetStyleCell(sheetManu,2,2,startColumnTypes,count+2,sType)      #Set Border for row 2
            SetBorderAll(sheetManu,3,3,startColumnTypes,count+2,'medium')   #Set Border for row 3
            current_type=matModels[count][0]
            startColumnTypes=count+2
        if count+1==len(matModels):                 #Last Element
            sType.fill=styles.fills.PatternFill(start_color=colorRandom, end_color=colorRandom, fill_type='solid')
            sheetManu.merge_cells(start_row=2,start_column=startColumnTypes,end_row=2,end_column=count+2)
            sheetManu.cell(row=2, column=startColumnTypes, value=current_type).style=sType
        #Copy value of model, so set the style to cell
        sheetManu.cell(row=3, column=count+2, value=matModels[count][1]).style=sModel
        count=count+1
    #Cycle to copy Locations in Excel file
    count=0
    for x in matLocations:
        sheetManu.cell(row=count+5, column=1, value=matLocations[count][1]).style=sLocation
        count=count+1
    #Set style of General Cell
    SetStyleCell(sheetManu,5,count+4,2,lModels+1,sText)
    #Set style of row 4 - Totali
    SetStyleCell(sheetManu,4,4,2,lModels+1,sTotali)
    # Adjust columns width - by column 1 (Locations)
    rng=CreateRangeCell(sheetManu,1,1,2,sheetManu.max_row)
    width_col(sheetManu,rng)
    # Adjust columns width - by column 2 to end (type)
    rng=CreateRangeCell(sheetManu,2,sheetManu.max_column,2,2)
    width_col(sheetManu,rng)
    return sheetManu

#FUNZIONI CELLE - RANGE
def CreateRangeCell(ws,col_min,col_max,row_min,row_max):
    arrCell=[]
    i=0
    arrCell=[[0]*1]
    r=row_min
    while r<=row_max:
        c=col_min
        while c<=col_max:
            if i!=0:
                arrCell.append([0]*1)
            arrCell[i]=ws.cell(row=r,column=c) 
            c=c+1
            i=i+1
        r=r+1
    return arrCell

def width_col(ws,arrCell):
    dims = {}
    for cell in arrCell:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = (value+2)*1.4


