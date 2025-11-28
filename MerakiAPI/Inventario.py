import time,sys,requests,json,openpyxl,pandas as pd
from openpyxl import load_workbook,Workbook

def get_invetario(URL,APIKEY,orgID,xls_path_inv):
    swListTot={"Name":[],"Serial":[],"Model":[], "IP":[], "NetwName":[], "portID":[], "Enabled":[], "Status":[], "PortName":[], "PoeEnable":[]}
    df_swListTot=pd.DataFrame([])
    timestr = time.strftime("%Y%m%d-%H")
    xls_path_inv_by_org = xls_path_inv+ timestr+"-TEC.xlsx"
    wb = openpyxl.Workbook()
    wb.save(filename = xls_path_inv_by_org)
    wb = openpyxl.load_workbook(xls_path_inv_by_org)
    CreateSheet(wb,'Inventory',xls_path_inv_by_org)
    ws = wb.active
    ws=wb['Inventory']

    # - 1 versione Get Inventory
    networkList = getNetworksID(URL,APIKEY,orgID)
    for ntwID in networkList:
        swList=getMyInventory(URL,APIKEY,ntwID,orgID)
        df_swList=pd.DataFrame(swList)
        df_swListTot=pd.concat([df_swListTot,df_swList])
        df_swListTot.to_excel(xls_path_inv_by_org,sheet_name=ws.title,index=False)


def get_invetarioByNetwork(URL,APIKEY,orgID,xls_path_inv):
    swListTot={"Name":[],"Serial":[],"Model":[], "IP":[], "NetwName":[], "portID":[], "Enabled":[], "Status":[], "PortName":[], "PoeEnable":[]}
    df_swListTot=pd.DataFrame([])
    timestr = time.strftime("%Y%m%d-%H")
    #Richiedo la network desiderata
    ntw=getNetworksID_Name(URL,APIKEY,orgID)
    ntwName=AskNetworkNameToUser(ntw)
    #Creazione file XLS
    xls_path_inv_by_ntw = xls_path_inv+ timestr+"-"+ntwName+".xlsx"
    wb = openpyxl.Workbook()
    wb.save(filename = xls_path_inv_by_ntw)
    wb = openpyxl.load_workbook(xls_path_inv_by_ntw)
    CreateSheet(wb,'Inventory',xls_path_inv_by_ntw)
    ws = wb.active
    ws=wb['Inventory']

    # - 1 versione Get Inventory

    ntwID=searchNetwIDByNetwName(ntwName,URL,APIKEY,orgID)
    
    swList=getMyInventory(URL,APIKEY,ntwID,orgID)
    df_swList=pd.DataFrame(swList)
    df_swListTot=pd.concat([df_swListTot,df_swList])
    df_swListTot.to_excel(xls_path_inv_by_ntw,sheet_name=ws.title,index=False)

def AskNetworkNameToUser(ntw):
    #print("Selezionare Network")
    print(*ntw["Name"], sep = "\n")
    ntwName=input("Digitare la network desiderata: ")
    return ntwName

def ReadDataExc(ws,row,col):
    data=""
    data=ws.cell(row=row,column=col)
    print(f"{data.value}")

def CreateSheet(wb,sheetName,save_path):
    found=0
    for sh_name in wb:
        if sheetName==sh_name.title:
            found=1
    if found==0:
        wb.create_sheet(sheetName)
    wb.save(save_path)

def getOrgID():
    """
    Query Meraki API for which Organizations we have access to & return Org ID
    """
    queryURL = URL + "/organizations"
    response = requests.get(queryURL, headers=APIKEY)
    orgID = json.loads(response.text)[0]["id"]
    return orgID

def getNetworksID(URL,APIKEY,orgID):
    """
    Use Organization ID to pull list of networks we have access to
    """
    queryURL = URL + f"/organizations/{orgID}/networks"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    networkID = []
    for network in data:
        networkID.append(network["id"])
    return networkID

def getNetworksID_Name(URL,APIKEY,orgID):
    """
    Use Organization ID to pull list of networks we have access to
    """
    queryURL = URL + f"/organizations/{orgID}/networks"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    network = {"ID":[],"Name":[]};
    for ntw in data:
        network["ID"].append(ntw["id"])
        network["Name"].append(ntw["name"])
    return network

def getClients(orgID, networkList):
    """
    Query clients for each network, return client list
    """
    clientCount = {}
    total = 0
    # Query Parameters: Return up to 100 devices seen in the past 43,200 seconds (30 days)
    q = {"perPage": "100",
         "timespan": "43200"}
    for network in networkList:
        # Query clients for each network
        queryURL = URL + f"/networks/{network}/clients"
        response = requests.get(queryURL, params=q, headers=APIKEY)
        data = json.loads(response.text)
        # Grab client OS from each device & append to clientCount dictionary
        for client in data:
            try:
                clientCount[client["os"]] += 1
            except KeyError:
                clientCount[client["os"]] = 1
            except TypeError:
                continue
            total += 1
    # Append final count of all devices & return dict
    clientCount["Total Devices"] = total
    return clientCount

def printReport(clientOS):
    """
    Print final output to terminal
    """
    print("Count of clients by operating system:")
    for OS in clientOS:
        print(f"{OS}: {clientOS[OS]}")

def ReturnSwitchByOrg(orgID):
    """
    Query per restituire tutti gli switch di un organization
    """
    swCount = {}
    total = 0
    for network in networkList:
        # Query clients for each network
        queryURL = URL + f"/networks/{network}/clients"
        response = requests.get(queryURL, params=q, headers=APIKEY)
        data = json.loads(response.text)
    return total

def getInventorySwitch(orgID):
    switchDict = {}
    queryURL = URL + f"/organizations/{orgID}/inventory/devices"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    swCount=0
    inventoryList = []
    #switches = [device  if device['model'][:2] in ('MS') and device['networkId'] is not None]
    for device in data:
        if device['model'][:2] in ('MS'):    
            try:
                switchDict[device['model']] +=1
            except KeyError:
                switchDict[device['model']] = 1
            except TypeError:
                continue
            swCount=swCount+1
    print(f"{swCount}")
    switchDict["Total Devices"] = swCount
    return switchDict

def getInventorySwitchByNetwork(ntwID):
    switchDict = {}
    queryURL = URL + f"/networks/{ntwID}/devices"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    inventoryList = {"Name":[],"Serial":[],"Model":[], "IP":[]};
    for device in data:
        if device['model'][:2] in ('MS'):
            inventoryList["Name"].append(device["name"])
            inventoryList["Serial"].append(device["serial"])
            inventoryList["Model"].append(device["model"])
            inventoryList["IP"].append(device["lanIp"])
    #print(f"{swCount}")
    return inventoryList

def getMyInventory(URL,APIKEY,ntwID,orgID):
    switchDict = {}
    ntwName=""
    #Get NetworkNames
    queryURL = URL + f"/organizations/{orgID}/networks"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    data = [x for x in data if x['id'] == ntwID]
    ntwName=data[0]['name']
    #Get Inventory devices
    queryURL = URL + f"/networks/{ntwID}/devices"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    inventoryList = {"Name":[],"Serial":[],"Model":[], "IP":[], "NetwName":[], "portID":[], "Enabled":[], "Status":[], "PortName":[], "PoeEnable":[]};
    for device in data:
        if device['model'][:2] in ('MS'):
            #Get Switchports by Serial
            #Queery per porte Abilitate ma DOWN - 2592000=1 Mese
            queryURL = URL + f"/devices/"+device["serial"]+"/switch/ports/statuses?timespan=2592000"
            response = requests.get(queryURL, headers=APIKEY)
            sw_port = json.loads(response.text)
            for x in sw_port:
                # Controllo che la porta sia abilitata ma in stato DOWN
                #trafic=x['usageInKb']['total']
                if (x['enabled'] == True) and (x['status'] in ('Disconnected')) and (x['usageInKb']['total']==0):
                    # Verifico che le porte non siano UPLINK (ovvero scarto porte > 24 per switch a 24 Porte e porte > 48 per switch 48 porte)
                    #test -- inizio
                    #portid=int(x['portId'])
                    #if int(x['portId'])<25:
                    #    a=0
                    #else:
                    #    a=1
                    #test -- fine
                    if (('-24') in device['model'] and int(x['portId'])<25) or (('-48') in device['model'] and int(x['portId'])<49):
                        queryPortDet = URL + f"/devices/"+device["serial"]+"/switch/ports/"+x['portId']
                        response = requests.get(queryPortDet, headers=APIKEY)
                        sw_port_det = json.loads(response.text)
                        # Verifico che nella description non ci sia scritto UPLINK
                        portname=sw_port_det['name']
                        if ((sw_port_det['name'] is not None and (('UPLINK') not in sw_port_det['name'])) or sw_port_det['name'] is None): 
                            inventoryList["Name"].append(device["name"])
                            inventoryList["Serial"].append(device["serial"])
                            inventoryList["Model"].append(device["model"])
                            inventoryList["IP"].append(device["lanIp"])
                            inventoryList["NetwName"].append(ntwName)
                            inventoryList["portID"].append(x['portId'])
                            inventoryList["Enabled"].append(x['enabled'])
                            inventoryList["Status"].append(x['status'])
                            inventoryList["PortName"].append(sw_port_det['name'])
                            inventoryList["PoeEnable"].append(sw_port_det['poeEnabled'])
    return inventoryList

def getInventorySwitchByNetworkGrpByModel(ntwID):
    switchDict = {}
    queryURL = URL + f"/networks/{ntwID}/devices"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    swCount=0
    inventoryList = []
    for device in data:
        if device['model'][:2] in ('MS'):    
            try:
                switchDict[device['model']] +=1
            except KeyError:
                switchDict[device['model']] = 1
            except TypeError:
                continue
            swCount=swCount+1
    #print(f"{swCount}")
    switchDict["Total Devices"] = swCount
    return switchDict

def searchNetwNameByID(ntwID):
    queryURL = URL + f"/organizations/{orgID}/networks"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    ntwName = ""
    for network in data:
        networkID=network["id"]
        if networkID==ntwID: 
            ntwName=(network["name"])
    return ntwName

def searchNetwIDByNetwName(ntwName,URL,APIKEY,orgID):
    queryURL = URL + f"/organizations/{orgID}/networks"
    response = requests.get(queryURL, headers=APIKEY)
    data = json.loads(response.text)
    ntwID = ""
    for network in data:
        networkName=network["name"]
        if networkName==ntwName: 
            ntwID=(network["id"])
    return ntwID

def PrintResult (var):
    for x in var:
        print(f"{x}")

def printReport(var):
    """
    Print final output to terminal
    """
    #print("COUNT GROUP BY")
    for grp in var:
        print(f"{grp}: {var[grp]}")

def commentUseful():
    #ReadDataExc(ws,3,1)
    #orgID = getOrgID()
    #print(f"{orgID}")
    # - 1 versione Get Inventory  ---INIZIO
    #networkList = getNetworksID(orgID)
    #for ntwID in networkList:
    #    swList=getInventorySwitchByNetwork(ntwID)
    #    df_swList=pd.DataFrame(swList)
    #    df_swListTot=pd.concat([df_swListTot,df_swList])
    #df_swListTot.to_excel(xls_path,sheet_name=ws.title,index=False,)
    # - 1 versione Get Inventory ---FINE

    #    swListTot["Name"].append(swList["Name"])
    #    swListTot["Serial"].append(swList["Serial"])
    #    swListTot["Model"].append(swList["Model"])
    #    swListTot["IP"].append(swList["IP"])
    #pd.DataFrame({'date' : dict_dates.keys() , 'date_value' : dict_dates.values() })
    #df=pd.DataFrame(list(swListTot.items()), columns=['Name', 'Serial', 'Model', 'IP'])
    #df.to_excel(xls_path,sheet_name=ws.title,index=False)
    #ws["A1"] = "Name"
    #ws["B1"] = "Serial"
    #ws["C1"] = "Model"
    #ws["D1"] = "IP"
    #for row, (name, serial, model, ip) in enumerate(swListTot["Name"], swListTot["Serial"],swListTot["Model"],swListTot["IP"]):
    #    ws [f"A{row}"] = name
    #    ws [f"B{row}"] = serial
    #    ws [f"C{row}"] = model
    #    ws [f"D{row}"] = ip
    #xls_path_inv_by_org = xls_path_inv+ timestr+"-TEC.xlsx"
    #wb.save(filename = xls_path_inv_by_org)
    #    ntwName=searchNetwNameByID(ntwID)
    #    print(f"{ntwName}")
    #    printReport(swList)
    #swList=getInventorySwitch(orgID)
    #inventoryList=getInventorySwitch(orgID)
    #PrintResult(inventoryList)
    #printReport(swList)
    #clientOS = getClients(orgID, networkList)
    #printReport(clientOS)
    debug=0
