import pandas as pd
import os
from openpyxl import styles
from openpyxl import Workbook, worksheet, styles, load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from Function.FuncDB import Func_PY_DB as FuncDB
from Function.FuncMatrix import Func_PY_Matrix as FuncMatrix 
from Function.FuncExcel import Func_PY_Excel as FuncExc
from config import InveManu_nameFile,ExOrange,ExBlack


def APP_GLPI_InveManu(entities,location,state):
    #Read Variable passed on form
    user_location=location
    user_entities=entities
    user_state=state

    #Set style Variable Excel
    sTitle1= NamedStyle(name="sTitle1")
    sTitle2= NamedStyle(name="sTitle2")
    sType= NamedStyle(name="sType")
    sModel= NamedStyle(name="sModel")
    sTotali= NamedStyle(name="sTotali")
    sLocation= NamedStyle(name="sLocation")
    sText=NamedStyle(name="sText")

    wb1=Workbook()

    #Crea File Excel vuoto
    #df=pd.DataFrame()
    #wb1=df.to_excel(InveManu_nameFile, index=False, engine='openpyxl')

    #Set Stile Foglio Excel - sTitle1 - style to Title (big size)
    sTitle1=FuncExc.SetStyleExcel_Font(sTitle1,20,True)
    sTitle1=FuncExc.SetStyleExcel_Alignement(sTitle1,'center','center')
    sTitle1=FuncExc.SetStyleExcel_PatternFill(sTitle1,ExOrange,ExOrange,'solid')
    sTitle1=FuncExc.SetStyleExcel_Border(sTitle1,'medium',ExBlack,'medium')
    #style to 2nd level title (more small than Title)
    sTitle2=FuncExc.SetStyleExcel_Font(sTitle2,14,True)
    sTitle2=FuncExc.SetStyleExcel_Alignement(sTitle2,'center','center')
    sTitle2=FuncExc.SetStyleExcel_PatternFill(sTitle2,ExOrange,ExOrange,'solid')
    sTitle2=FuncExc.SetStyleExcel_Border(sTitle2,'medium',ExBlack,'medium')
    #Style to row 2 Type
    sType=FuncExc.SetStyleExcel_Font(sType,14,True)
    sType=FuncExc.SetStyleExcel_Alignement(sType,'center','center')
    sType=FuncExc.SetStyleExcel_Border(sType,'medium',ExBlack,'medium')
    #Style to row 3 Models
    sModel=FuncExc.SetStyleExcel_Font(sModel,14,True)
    sModel=FuncExc.SetStyleExcel_Alignement(sModel,'center','center')
    sModel=FuncExc.SetStyleExcel_Rotation(sModel,90)
    sModel=FuncExc.SetStyleExcel_Border(sModel,'thin',ExBlack,'thin')
    #Style to row 4 'Totali'
    sTotali=FuncExc.SetStyleExcel_Font(sTotali,14,True)
    sTotali=FuncExc.SetStyleExcel_Alignement(sTotali,'center','center')
    sTotali=FuncExc.SetStyleExcel_Border(sTotali,'medium',ExBlack,'medium')
    #Style of column Location Name
    sLocation=FuncExc.SetStyleExcel_Font(sLocation,12,True)
    sLocation=FuncExc.SetStyleExcel_Alignement(sLocation,'center','center')
    sLocation=FuncExc.SetStyleExcel_PatternFill(sLocation,ExOrange,ExOrange,'solid')
    sLocation=FuncExc.SetStyleExcel_Border(sLocation,'thin',ExBlack,'thin')
    #Style of other cell
    sText=FuncExc.SetStyleExcel_Font(sText,12,False)
    sText=FuncExc.SetStyleExcel_Alignement(sText,'center','center')
    sText=FuncExc.SetStyleExcel_Border(sText,'thin',ExBlack,'thin')

    #Extract array from query on GLPI DB
    arrEntities=FuncDB.CopiaCampiDB("SELECT * FROM glpi_entities",1,0)                                              #Colonna 0 = id | colonna 1 = name
    arrLocations=FuncDB.CopiaCampiDB("SELECT * FROM glpi_locations",3,0)                                            #Colonna 0 = id | colonna 3 = name
    arrNetweqTypes=FuncDB.CopiaCampiDB("SELECT * FROM glpi_networkequipmenttypes",1,0)                              #Colonna 0 = id | colonna 1 = name
    arrNetweqModels=FuncDB.CopiaCampiDB("SELECT * FROM glpi_networkequipmentmodels",1,0)                            #Colonna 0 = id | colonna 1 = name
    arrVendors=FuncDB.CopiaCampiDB("SELECT * FROM glpi_manufacturers",1,0)                                          #Colonna 0 = id | colonna 1 = name
    arrStates=FuncDB.CopiaCampiDB("SELECT * FROM glpi_states",1,0)                                                  #Colonna 0 = id | colonna 1 = name    
    arrIPAddresses=FuncDB.CopiaCampiDB("SELECT * FROM glpi_ipaddresses",5,12)                                       #Colonna 5 = id | colonna 12 = name
    arrMACAddresses=FuncDB.CopiaCampiDB("SELECT * FROM glpi_networkports",8,1)                                      #Colonna 1 = items_id | colonna 8 = mac
    arrPropHW=FuncDB.CopiaCampiDB("SELECT * FROM glpi_plugin_fields_proprietariohwfielddropdowns",1,0)              #da Plugin - Proprietario HW 
    arrMaintSuppl=FuncDB.CopiaCampiDB("SELECT * FROM glpi_plugin_fields_maintenancesupplierfielddropdowns",1,0)     #da Plugin - Maintenance Supplier
    arrGenericPlugin_Model=FuncDB.CopiaCampiDB("SELECT * FROM glpi_plugin_genericobject_oxomodels",1,0)             #da Plugin - Model Plugin Generci (es. per OXO)
    arrGenericPlugin_Type=FuncDB.CopiaCampiDB("SELECT * FROM glpi_plugin_genericobject_types",4,0)                  #da Plugin - Type Plugin Generci (es. per OXO)
    arrPhoneTypes=FuncDB.CopiaCampiDB("SELECT * FROM glpi_phonetypes",1,0)                                          #Array Tipo per Phones
    arrPhoneModels=FuncDB.CopiaCampiDB("SELECT * FROM glpi_phonemodels",1,0)                                        #Array Modello per Phones

    #ATTENZIONE!!!!!-------ATTUALMENTE DISPONIBILE SOLO VERSIONE CON SCELTA ENTITA'------------
    #if((user_location==['empty']) and (user_entities!=[])):       #user selected the entities
    if(user_entities!=[]):       #user selected the entities
        #cod_ent=int(user_entities[0])       #read Value selected from user and convert it to a Integer
        arrEntSelected=FuncMatrix.ConvertArray(user_entities,arrEntities)
        queryAdd=FuncDB.CreateAddQuery(arrEntSelected)
        #Find Name from code Entities
        #ent=FindElemMatrix(arrEntities,cod_ent,0,1)
        #query1="SELECT * FROM glpi_networkequipments INNER JOIN glpi_entities ON glpi_networkequipments.entities_id=glpi_entities.id WHERE glpi_entities.name LIKE '%" + ent + "%' AND glpi_networkequipments.is_template='0'"

        #####################Query per recuperare dati NETWORK - da Tabella glpi_networkequipments ####################
        query_NTWEquipment="SELECT * FROM glpi_networkequipments \
        INNER JOIN glpi_entities ON glpi_networkequipments.entities_id=glpi_entities.id \
        WHERE ("+queryAdd+") AND glpi_networkequipments.is_template='0'"
#       Conversione N.colonne=Description 0=id, 13=locations_id, 3=name, 23=states_id, 15=networkequipmenttypes_id, 17=manufacturers_id, 16=networkequipmentmodels_id, 1=entities_id)
        index_id=0
        index_locations_id=13
        index_name=3
        index_states_id=23
        index_networkequipmenttypes_id=15
        index_manufacturers_id=17
        index_networkequipmentmodels_id=16
        index_entities_id=1
        #create matrix with all info necessary
        arrRisQuery_NTWEquipment=FuncDB.CopiaRisQuery_14_filed(query_NTWEquipment,
                                                               index_id,index_locations_id,index_name,index_states_id,index_networkequipmenttypes_id,index_manufacturers_id,index_networkequipmentmodels_id,index_entities_id,
                                                               arrLocations,arrStates,arrNetweqTypes,arrVendors,arrNetweqModels,arrEntities)
        #####################Query per recuperare dati CENTRALI TELEFONICHE - da Plugin OXO ####################
        query_OXO="SELECT * FROM glpi_plugin_genericobject_oxos \
        INNER JOIN glpi_entities ON glpi_plugin_genericobject_oxos.entities_id=glpi_entities.id \
        WHERE ("+queryAdd+") AND glpi_plugin_genericobject_oxos.is_template='0'"
#       Conversione N.colonne=Description 0=id, 9=locations_id, 6=name, 10=states_id, networkequipmenttypes_id-99-OXO, 12=manufacturers_id, 13=PlgGen=networkequipmentmodels_id, 4=entities_id)
        index_id=0
        index_locations_id=9
        index_name=6
        index_states_id=10
        index_networkequipmenttypes_id=99  # Valore Arbitrario per poterlo forzare in OXO - è dato dal fatto che in query_FONIA sia specificato tabella OXO
        index_manufacturers_id=12
        index_networkequipmentmodels_id=13
        index_entities_id=4
        #create matrix with all info necessary 
        arrRisQuery_OXO=FuncDB.CopiaRisQuery_14_filed(query_OXO,
                                                      index_id,index_locations_id,index_name,index_states_id,index_networkequipmenttypes_id,index_manufacturers_id,index_networkequipmentmodels_id,index_entities_id,
                                                      arrLocations,arrStates,arrGenericPlugin_Type,arrVendors,arrGenericPlugin_Model,arrEntities)
        #####################Query per recuperare dati FONIA (Telefoni+IBS+Schede) - da Plugin OXO ####################
        query_FONIA="SELECT * FROM glpi_phones \
        INNER JOIN glpi_entities ON glpi_phones.entities_id=glpi_entities.id \
        WHERE ("+queryAdd+") AND glpi_phones.is_template='0'"
#       Conversione N.colonne=Description 0=id, 11=locations_id, 2=name, 26=states_id, 12=Phonetypes_id, 19=manufacturers_id, 13=PhoneModel_id_id, 4=entities_id)
        index_id=0
        index_locations_id=11
        index_name=2
        index_states_id=26
        index_networkequipmenttypes_id=12  
        index_manufacturers_id=19
        index_networkequipmentmodels_id=13
        index_entities_id=1
        #create matrix with all info necessary 
        arrRisQuery_FONIA=FuncDB.CopiaRisQuery_14_filed(query_FONIA,
                                                      index_id,index_locations_id,index_name,index_states_id,index_networkequipmenttypes_id,index_manufacturers_id,index_networkequipmentmodels_id,index_entities_id,
                                                      arrLocations,arrStates,arrPhoneTypes,arrVendors,arrPhoneModels,arrEntities)
#       Unisco tutti gli array
        arrRisQuery= arrRisQuery_NTWEquipment + arrRisQuery_OXO + arrRisQuery_FONIA
#       Solo per TEST
        #arrRisQuery= arrRisQuery_NTWEquipment


    #LL=len(arrRisQuery)
    #OLD- DA VERIFICARE E CAPIRE ----arrStateName=FuncMatrix.ConvertArray(user_state,arrStates)                     #extract Name from state selected from user
    arrStateName=user_state #Sostituisce la precedente da verificare e capire
    matrixModels=FuncMatrix.ExtractMatrix_2_index(arrRisQuery,7,11)                #extract type_name and model_name from arrRisQuery
    matrixModelsUnique=FuncMatrix.DeleteDuplicateMatrix(matrixModels,1,0)          #delete duplicated element from matrixModels
    matrixLocations=FuncMatrix.ExtractMatrix_2_index(arrRisQuery,1,2)              #extract locations_id and locations_name from arrRisQuery
    matrixLocationsUnique=FuncMatrix.DeleteDuplicateMatrix(matrixLocations,1,0)    #delete duplicated element from matrixLocations
    #create sheet with correctb style
    sheetManu=FuncExc.createSheetManu(wb1,matrixModelsUnique,matrixLocationsUnique,arrRisQuery,sTitle1,sTitle2,sType,sModel,sTotali,sLocation,sText)
    #exc_save_filename="\\Manu_InveTEST1.xlsx"
    #saveExcel(exc_path,exc_save_filename,wb1)
    #count the number of device
    FuncExc.countDevice(sheetManu,arrRisQuery,arrStateName)

    FuncExc.SumDevice(sheetManu)
    breakpoint=0 #solo per mettere i breakpoint, Variable useless
    #operation on GLPI DB
    #myDB=funzioni.connetti_db()                | TEST INSTRUCTION
    #query="SELECT * FROM glpi_apiclients"      | TEST INSTRUCTION
    #output=funzioni.exec_query(myDB,query)     | TEST INSTRUCTION

    #operation on EXC file
    
    wb1.save(filename=InveManu_nameFile)
    #funzioniOS.saveExcel(exc_path,exc_save_filename,wb1)

    #Restituisci percorso del file Excel
    return InveManu_nameFile
    #Extract array from query on GLPI DB
    #arrEntities=Func_PY_DB.CopiaCampiDB("SELECT * FROM glpi_entities",1,0)